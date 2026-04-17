from __future__ import annotations

import uuid
from datetime import date, timedelta
from io import BytesIO
from unittest.mock import MagicMock

import openpyxl
import pytest

from app.services.export_service import export_route_group


# ── helpers ──────────────────────────────────────────────────────────────────

def make_route_group() -> MagicMock:
    rg = MagicMock()
    rg.id = uuid.uuid4()
    rg.name = "Test Group"
    return rg


def make_result(
    origin: str = "AMD",
    destination: str = "SGN",
    depart_date: date | None = None,
    price: float = 200.0,
    airline: str = "VJ",
    provider: str = "serpapi",
    stops: int | None = 0,
    duration_minutes: int | None = 120,
    currency: str = "USD",
) -> MagicMock:
    r = MagicMock()
    r.origin = origin
    r.destination = destination
    r.depart_date = depart_date or (date.today() + timedelta(days=1))
    r.price = price
    r.airline = airline
    r.provider = provider
    r.stops = stops
    r.duration_minutes = duration_minutes
    r.currency = currency
    return r


# ── tests ─────────────────────────────────────────────────────────────────────

def test_export_creates_all_results_sheet() -> None:
    rg = make_route_group()
    excel_bytes = export_route_group(rg, [make_result()])
    wb = openpyxl.load_workbook(BytesIO(excel_bytes))
    assert "All Results" in wb.sheetnames
    assert len(wb.sheetnames) == 1


def test_export_empty_results_produces_header_only() -> None:
    rg = make_route_group()
    excel_bytes = export_route_group(rg, [])
    wb = openpyxl.load_workbook(BytesIO(excel_bytes))
    ws = wb["All Results"]
    # Only header row, no data rows
    assert ws.max_row == 1


def test_export_has_correct_headers() -> None:
    rg = make_route_group()
    excel_bytes = export_route_group(rg, [])
    wb = openpyxl.load_workbook(BytesIO(excel_bytes))
    ws = wb["All Results"]
    assert ws.cell(1, 1).value == "Date"
    assert ws.cell(1, 2).value == "Dep Airport"
    assert ws.cell(1, 3).value == "Arr Airport"
    assert ws.cell(1, 4).value == "Airline"
    assert ws.cell(1, 5).value == "Price"
    assert ws.cell(1, 6).value == "Currency"
    assert ws.cell(1, 7).value == "Stops"
    assert ws.cell(1, 8).value == "Duration (min)"
    assert ws.cell(1, 9).value == "Provider"


def test_export_prices_are_integers() -> None:
    rg = make_route_group()
    excel_bytes = export_route_group(rg, [make_result(price=199.75)])
    wb = openpyxl.load_workbook(BytesIO(excel_bytes))
    ws = wb["All Results"]
    assert ws.cell(2, 5).value == 200


def test_export_sorted_by_date_then_price() -> None:
    rg = make_route_group()
    today = date.today()
    r1 = make_result(depart_date=today + timedelta(days=2), price=300.0)
    r2 = make_result(depart_date=today + timedelta(days=1), price=500.0)
    r3 = make_result(depart_date=today + timedelta(days=1), price=200.0)

    excel_bytes = export_route_group(rg, [r1, r2, r3])
    wb = openpyxl.load_workbook(BytesIO(excel_bytes))
    ws = wb["All Results"]

    # Row 2: earliest date, cheapest price first
    assert ws.cell(2, 5).value == 200
    assert ws.cell(3, 5).value == 500
    assert ws.cell(4, 5).value == 300


def test_export_none_stops_shows_dash() -> None:
    rg = make_route_group()
    excel_bytes = export_route_group(rg, [make_result(stops=None)])
    wb = openpyxl.load_workbook(BytesIO(excel_bytes))
    ws = wb["All Results"]
    assert ws.cell(2, 7).value == "-"


def test_export_zero_duration_shows_dash() -> None:
    rg = make_route_group()
    excel_bytes = export_route_group(rg, [make_result(duration_minutes=0)])
    wb = openpyxl.load_workbook(BytesIO(excel_bytes))
    ws = wb["All Results"]
    assert ws.cell(2, 8).value == "-"


def test_export_provider_written_to_column_9() -> None:
    rg = make_route_group()
    excel_bytes = export_route_group(rg, [make_result(provider="serpapi")])
    wb = openpyxl.load_workbook(BytesIO(excel_bytes))
    ws = wb["All Results"]
    assert ws.cell(2, 9).value == "serpapi"
