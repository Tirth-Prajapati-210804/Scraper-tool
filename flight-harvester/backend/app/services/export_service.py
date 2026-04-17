from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from app.models.all_flight_result import AllFlightResult
from app.models.route_group import RouteGroup

_MAIN_HEADERS = ["Date", "Dep Airport", "Arrivel Airport", "Night ", "Airline", "Flight Price"]
_SPECIAL_HEADERS_4 = ["Date", "Dep Airport", "Arrivel Airport", "Flight Price"]
_SPECIAL_HEADERS_6 = ["Date", "Dep Airport", "Arrivel Airport", "Night ", "Airline", "Flight Price"]


def export_route_group(
    route_group: RouteGroup,
    all_results: list[AllFlightResult],
) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    # Build cheapest-per-(origin, date) lookup across all destinations
    cheapest: dict[tuple[str, object], AllFlightResult] = {}
    for r in all_results:
        key = (r.origin, r.depart_date)
        if key not in cheapest or r.price < cheapest[key].price:
            cheapest[key] = r

    # Full sorted date range from collected data
    all_dates = sorted({r.depart_date for r in all_results})

    # ── Main sheets — one per origin ─────────────────────────────────────────
    for origin, sheet_name in route_group.sheet_name_map.items():
        ws = wb.create_sheet(title=sheet_name)
        _write_header_row(ws, _MAIN_HEADERS)

        for row_idx, d in enumerate(all_dates, start=2):
            result = cheapest.get((origin, d))
            ws.cell(row=row_idx, column=1, value=d).number_format = "YYYY-MM-DD"
            ws.cell(row=row_idx, column=2, value=origin.strip())
            ws.cell(row=row_idx, column=3, value=route_group.destination_label)
            ws.cell(row=row_idx, column=4, value=route_group.nights)
            ws.cell(row=row_idx, column=5, value=result.airline if result else None)
            ws.cell(row=row_idx, column=6, value=int(round(float(result.price))) if result else None)

        _autosize_columns(ws)

    # ── Special sheets ────────────────────────────────────────────────────────
    for spec in route_group.special_sheets:
        spec_origin = spec["origin"]
        spec_dests = set(spec["destinations"])
        columns = spec.get("columns", 6)

        # Cheapest per date across this special sheet's destinations
        spec_cheapest: dict[object, AllFlightResult] = {}
        for r in all_results:
            if r.origin == spec_origin and r.destination in spec_dests:
                if r.depart_date not in spec_cheapest or r.price < spec_cheapest[r.depart_date].price:
                    spec_cheapest[r.depart_date] = r

        spec_dates = sorted(spec_cheapest.keys()) if spec_cheapest else all_dates

        ws = wb.create_sheet(title=spec["name"])
        headers = _SPECIAL_HEADERS_4 if columns == 4 else _SPECIAL_HEADERS_6
        _write_header_row(ws, headers)

        for row_idx, d in enumerate(spec_dates, start=2):
            result = spec_cheapest.get(d)
            ws.cell(row=row_idx, column=1, value=d).number_format = "YYYY-MM-DD"
            ws.cell(row=row_idx, column=2, value=spec_origin)
            ws.cell(row=row_idx, column=3, value=spec["destination_label"])
            if columns == 4:
                ws.cell(row=row_idx, column=4, value=int(round(float(result.price))) if result else None)
            else:
                ws.cell(row=row_idx, column=4, value=route_group.nights)
                ws.cell(row=row_idx, column=5, value=result.airline if result else None)
                ws.cell(row=row_idx, column=6, value=int(round(float(result.price))) if result else None)

        _autosize_columns(ws)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


def _write_header_row(ws: object, headers: list[str]) -> None:
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)  # type: ignore[union-attr]
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")


def _autosize_columns(ws: object) -> None:
    for col_cells in ws.columns:  # type: ignore[union-attr]
        max_length = max((len(str(c.value)) for c in col_cells if c.value is not None), default=0)
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = max_length + 3  # type: ignore[union-attr]
